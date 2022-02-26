from pandas.io.json import json_normalize

import json
import pandas as pd
import re
import sys
import os
import openpyxl
import wget
wrkbk = openpyxl.Workbook()
ws = wrkbk.active
print(str(sys.argv))
meta_list=[]
path = "/home/unknown/Desktop/iom_files/Gut_Database/demo"
def listToString(s):
    # initialize an empty string
    string = ""
    # traverse in the string
    for element in s:
        string += element
    # return string
    return string


def Union(lst1, lst2):
    final_list = list(set(lst1) | set(lst2))
    return final_list


def genotypic(filename):
    file=filename+".json"
    data = json.load(open(file))
    print(len(data))
    for x in data:
        keys=x.keys()
    geattr=list(keys)
    columnlen=len(geattr)
    for i in range(columnlen):
        ws.cell(row=1,column=i+1).value=geattr[i]
    for i in range(columnlen):
        for k in range(len(data)):
            attributes=geattr[i]
            print(data[k][attributes])
            ws.cell(row=k+2,column=i+1).value=data[k][attributes]
    for w in range(len(data)):
        url = 'https://www.ebi.ac.uk/biosamples/samples/'+data[w]['sample_accession']+'.json'
        wget.download(url)
        data_p = json.load(open(data[w]['sample_accession']+'.json'))
        df = json_normalize(data_p, record_path =['characteristics'])
        fields=df.values.tolist()
        for x in range(len(fields)):
            meta_list.append(listToString(fields[x]))
        with open('test.txt') as f:
            content = f.readlines()
        li = [x.strip() for x in content]
        meta = Union(li,meta_list)
        textfile = open("test.txt", "w")
        for element in meta:
            textfile.write(element + "\n")
        textfile.close() 
    for w in range(len(data)):
        data_p1 = json.load(open(data[w]['sample_accession']+'.json'))
        df = json_normalize(data_p1, record_path =['characteristics'])

        with open('test.txt') as f:
            content = f.readlines()
        li = [x.strip() for x in content]
        for k in range(len(li)):
            ws.cell(row=1,column=columnlen+1+k).value=li[k]

        for k in range(len(li)):
            variables=str(li[k])
            try:
                print(variables+"="+data_p1['characteristics'][variables][0]['text'])
                ws.cell(row=w+2,column=columnlen+1+k).value=data_p1['characteristics'][variables][0]['text']
            except:
                print(variables+"= NA")
                ws.cell(row=w+2,column=columnlen+1+k).value="NA"



def check():
    if(sys.argv[1]):
        all_projects=os.listdir(path)
        print(type(all_projects))
        if sys.argv[1] not in all_projects:
           print('Creating a directory')
           new_directory=os.path.join(path,sys.argv[1])
           print(new_directory)
           os.mkdir(new_directory)
           os.chdir(new_directory)
           os.system("wget -O "+sys.argv[1]+".json " +'"https://www.ebi.ac.uk/ena/portal/api/filereport?accession='+sys.argv[1]+'&result=read_run&fields=study_accession,sample_accession,experiment_accession,run_accession,tax_id,scientific_name,fastq_ftp,submitted_ftp,sra_ftp&format=json&download=true&limit=0"')
           os.system("touch test.txt")
           genotypic(sys.argv[1])

           wrkbk.save(sys.argv[1]+'.xlsx')

        else:
           print("in else")
check()

